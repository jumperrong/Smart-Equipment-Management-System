"""结构化表单记录 API：基于模板填写 → 保存/提交/导出；关联工艺文件列表条目。"""
from __future__ import annotations

import csv
import io
import json
import os
import uuid
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.models import (
    FormRecord, FormRecordValue, FormTemplate,
    ProcessDocument, Equipment, User,
)
from app.schemas import FormRecordCreate, FormRecordOut, FormRecordUpdate
from app.services.user_service import get_current_user
from app.services.permission_service import require_permission
from app.services.form_template_service import (
    auto_generate_record_title,
    get_template_or_404,
    sorted_fields_from_template,
    validate_required_fields,
)

router = APIRouter(prefix="/form-records", tags=["表单记录"])


def _record_to_out(record: FormRecord, db: Session, include_values: bool = True) -> FormRecordOut:
    """ORM 实例 → Pydantic 输出；补展示辅助字段。"""
    tpl = db.query(FormTemplate).filter(FormTemplate.id == record.template_id).first()
    eq = None
    if record.equipment_id:
        eq = db.query(Equipment).filter(Equipment.id == record.equipment_id).first()
    values: list[dict] = []
    if include_values:
        for v in record.values:
            values.append({"field_key": v.field_key, "field_value": v.field_value})
    return FormRecordOut(
        id=record.id,
        template_id=record.template_id,
        title=record.title,
        equipment_id=record.equipment_id,
        batch_no=record.batch_no,
        shift=record.shift,
        production_date=record.production_date,
        remark=record.remark,
        status=record.status,
        filled_by=record.filled_by,
        submitted_at=record.submitted_at,
        values=values,
        template_name=tpl.name if tpl else None,
        template_category=tpl.category if tpl else None,
        equipment_name=eq.name if eq else None,
        created_at=record.created_at,
        updated_at=record.updated_at,
    )


def _apply_values_to_record(db: Session, record: FormRecord, kv_list: list[dict], fields: list[dict]) -> None:
    """增量/全量写入 values。传入 [{field_key, field_value}] 列表。

    - 只接受 template.fields 中存在的 key，忽略未知 key
    - 字段 label 快照写入（当新建/模板改 label 后仍能回看）
    - 同 record 重复 key 走 upsert
    """
    fields_by_key = {f["key"]: f for f in fields}
    existing = {v.field_key: v for v in record.values}
    for kv in kv_list or []:
        key = kv.get("field_key")
        if not key or key not in fields_by_key:
            continue
        val = kv.get("field_value")
        field = fields_by_key[key]
        label = field.get("label") or key
        if key in existing:
            existing[key].field_value = val
            existing[key].field_label_snapshot = label
        else:
            v = FormRecordValue(
                record_id=record.id,
                field_key=key,
                field_label_snapshot=label,
                field_value=val,
            )
            db.add(v)


# ============================================================
# 列表
# ============================================================

@router.get("", response_model=list[FormRecordOut])
def list_records(
    template_id: Optional[int] = Query(None),
    equipment_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None, description="草稿/已提交/已作废"),
    batch_no: Optional[str] = Query(None, description="批号模糊匹配"),
    shift: Optional[str] = Query(None),
    production_date_from: Optional[str] = Query(None),
    production_date_to: Optional[str] = Query(None),
    keyword: Optional[str] = Query(None, description="标题模糊匹配"),
    limit: int = Query(200, ge=1, le=1000),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    q = db.query(FormRecord)
    if template_id:
        q = q.filter(FormRecord.template_id == template_id)
    if equipment_id:
        q = q.filter(FormRecord.equipment_id == equipment_id)
    if status:
        q = q.filter(FormRecord.status == status)
    if batch_no:
        q = q.filter(FormRecord.batch_no.ilike(f"%{batch_no}%"))
    if shift:
        q = q.filter(FormRecord.shift == shift)
    if production_date_from:
        try:
            dt_from = datetime.fromisoformat(production_date_from.replace("Z", ""))
            q = q.filter(FormRecord.production_date >= dt_from)
        except ValueError:
            pass
    if production_date_to:
        try:
            dt_to = datetime.fromisoformat(production_date_to.replace("Z", ""))
            # 仅传日期(如 2026-08-20)时 fromisoformat 解析为当天 00:00，
            # 会漏掉当天 00:00 之后的记录；补到当天 23:59:59 兜底。
            if len(production_date_to) == 10:
                dt_to = dt_to.replace(hour=23, minute=59, second=59)
            q = q.filter(FormRecord.production_date <= dt_to)
        except ValueError:
            pass
    if keyword:
        q = q.filter(FormRecord.title.ilike(f"%{keyword}%"))
    records = q.order_by(FormRecord.created_at.desc()).limit(limit).all()
    # 列表默认不返回每个字段的值，避免过大
    return [_record_to_out(r, db, include_values=False) for r in records]


# ============================================================
# 创建：基于模板生成记录
# ============================================================

@router.post(
    "",
    response_model=FormRecordOut,
    dependencies=[Depends(require_permission("form_record.fill"))],
)
def create_record(
    payload: FormRecordCreate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    tpl = get_template_or_404(db, payload.template_id)
    if not tpl.is_active:
        raise HTTPException(status_code=400, detail=f"模板[{tpl.name}]已停用，不能用于生成新记录")

    # 若模板绑定了设备，允许用户传 equipment_id 覆盖或留空（空则继承模板的equipment_id）
    equipment_id = payload.equipment_id or tpl.equipment_id
    if equipment_id:
        eq = db.query(Equipment).filter(Equipment.id == equipment_id).first()
        if not eq:
            raise HTTPException(status_code=404, detail=f"机台 id={equipment_id} 不存在")

    # 自动生成标题
    prod_iso = payload.production_date.isoformat() if payload.production_date else None
    title = payload.title or auto_generate_record_title(tpl.name, payload.batch_no, prod_iso)

    status = "已提交" if payload.auto_submit else "草稿"
    submitted_at = datetime.utcnow() if payload.auto_submit else None

    record = FormRecord(
        template_id=tpl.id,
        title=title,
        equipment_id=equipment_id,
        batch_no=payload.batch_no,
        shift=payload.shift,
        production_date=payload.production_date,
        remark=payload.remark,
        status=status,
        filled_by=current_user.id,
        submitted_at=submitted_at,
    )
    db.add(record)
    db.flush()  # 获取 record.id

    fields = sorted_fields_from_template(tpl)
    # 初始化 values：先按 default_value 预填，再覆盖用户传入值
    initial_kv = []
    user_kv_by_key = {kv.field_key: kv.field_value for kv in (payload.values or [])}
    for f in fields:
        key = f["key"]
        v = user_kv_by_key.get(key, f.get("default_value"))
        initial_kv.append({"field_key": key, "field_value": v})
    _apply_values_to_record(db, record, initial_kv, fields)

    # 必填校验（仅 auto_submit=true 时）
    if payload.auto_submit:
        validate_required_fields(fields, {k["field_key"]: k["field_value"] for k in initial_kv})

    # 同步创建关联 ProcessDocument (category=record)，便于在工艺文件列表中看到
    if payload.link_process_doc:
        prod_date_str = (str(payload.production_date)[:10]) if payload.production_date else None
        group_id = uuid.uuid4().hex
        pdoc = ProcessDocument(
            equipment_id=equipment_id or (tpl.equipment_id or 0),
            category="record",
            doc_name=title,
            doc_type="BatchRecord",
            group_id=group_id,
            version_seq=1,
            is_latest=True,
            status="草稿",
            batch_no=payload.batch_no,
            shift=payload.shift,
            production_date=payload.production_date or None,
            stored_path="",  # 特殊标记：结构化表单记录无上传文件
            file_size=0,
            file_type="application/x-sems-form-record",
            description=tpl.description,
            uploaded_by=current_user.id,
            form_record_id=record.id,
        )
        # equipment_id=0 非法；若找不到可用 equipment_id 则报错并提示用户填写
        if not pdoc.equipment_id:
            raise HTTPException(status_code=400, detail="关联到工艺记录必须指定设备(equipment_id)，请选择机台后重试。")
        db.add(pdoc)

    db.commit()
    db.refresh(record)
    return _record_to_out(record, db, include_values=True)


# ============================================================
# 详情（含 values）
# ============================================================

@router.get("/{record_id}", response_model=FormRecordOut)
def get_record(
    record_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    rec = db.query(FormRecord).filter(FormRecord.id == record_id).first()
    if not rec:
        raise HTTPException(status_code=404, detail="表单记录不存在")
    return _record_to_out(rec, db, include_values=True)


# ============================================================
# 更新（编辑草稿/已提交均可；提交状态不阻止编辑以便后续更正）
# ============================================================

@router.put(
    "/{record_id}",
    response_model=FormRecordOut,
    dependencies=[Depends(require_permission("form_record.fill"))],
)
def update_record(
    record_id: int,
    payload: FormRecordUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    rec = db.query(FormRecord).filter(FormRecord.id == record_id).first()
    if not rec:
        raise HTTPException(status_code=404, detail="表单记录不存在")
    if rec.status == "已作废":
        raise HTTPException(status_code=400, detail="已作废的记录不允许编辑")
    # 文控审核锁定：已审核通过的记录禁止原地修改，只能通过 Amendment 附加修正
    if rec.audited:
        raise HTTPException(
            status_code=400,
            detail="该记录已文控审核锁定，禁止原地修改。如需更正，请通过「附加修正(Amendment)」留痕。"
        )

    tpl = get_template_or_404(db, rec.template_id)
    fields = sorted_fields_from_template(tpl)

    data = payload.model_dump(exclude_unset=True, exclude={"values"})
    for k, v in data.items():
        setattr(rec, k, v)

    if payload.values is not None:
        _apply_values_to_record(
            db, rec,
            [{"field_key": kv.field_key, "field_value": kv.field_value} for kv in payload.values],
            fields,
        )

    db.commit()
    db.refresh(rec)
    return _record_to_out(rec, db, include_values=True)


# ============================================================
# 提交：草稿→已提交，写入submitted_at + 必填校验
# ============================================================

@router.patch(
    "/{record_id}/submit",
    response_model=FormRecordOut,
    dependencies=[Depends(require_permission("form_record.fill"))],
)
def submit_record(
    record_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    rec = db.query(FormRecord).filter(FormRecord.id == record_id).first()
    if not rec:
        raise HTTPException(status_code=404, detail="表单记录不存在")
    if rec.status == "已提交":
        return _record_to_out(rec, db, include_values=True)
    if rec.status == "已作废":
        raise HTTPException(status_code=400, detail="已作废记录不可提交")
    tpl = get_template_or_404(db, rec.template_id)
    fields = sorted_fields_from_template(tpl)
    values_by_key = {v.field_key: v.field_value for v in rec.values}
    validate_required_fields(fields, values_by_key)
    rec.status = "已提交"
    rec.submitted_at = datetime.utcnow()
    if not rec.filled_by:
        rec.filled_by = current_user.id
    db.commit()
    db.refresh(rec)
    return _record_to_out(rec, db, include_values=True)


# ============================================================
# 作废
# ============================================================

@router.patch(
    "/{record_id}/void",
    response_model=FormRecordOut,
    dependencies=[Depends(require_permission("form_record.fill"))],
)
def void_record(
    record_id: int,
    db: Session = Depends(get_db),
):
    rec = db.query(FormRecord).filter(FormRecord.id == record_id).first()
    if not rec:
        raise HTTPException(status_code=404, detail="表单记录不存在")
    if rec.status == "已作废":
        return _record_to_out(rec, db, include_values=True)
    rec.status = "已作废"
    db.commit()
    db.refresh(rec)
    return _record_to_out(rec, db, include_values=True)


# ============================================================
# 删除
# ============================================================

@router.delete(
    "/{record_id}",
    dependencies=[Depends(require_permission("form_record.delete"))],
)
def delete_record(
    record_id: int,
    db: Session = Depends(get_db),
):
    rec = db.query(FormRecord).filter(FormRecord.id == record_id).first()
    if not rec:
        raise HTTPException(status_code=404, detail="表单记录不存在")
    # 解除关联 process_doc.form_record_id 以避免误删联动
    for pd in rec.process_documents:
        pd.form_record_id = None
    db.flush()
    db.delete(rec)
    db.commit()
    return {"ok": True, "removed": record_id}


# ============================================================
# 导出 JSON / CSV
# ============================================================

@router.get("/{record_id}/export/json")
def export_json(
    record_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    rec = db.query(FormRecord).filter(FormRecord.id == record_id).first()
    if not rec:
        raise HTTPException(status_code=404, detail="表单记录不存在")
    out = _record_to_out(rec, db, include_values=True).model_dump(mode="json")
    # 把 values 数组转成 {label: value} 更易读
    fields = sorted_fields_from_template(get_template_or_404(db, rec.template_id))
    label_by_key = {f["key"]: f.get("label") or f["key"] for f in fields}
    pretty = {}
    for kv in out.get("values", []):
        k = kv.get("field_key")
        pretty[label_by_key.get(k, k)] = kv.get("field_value")
    out.pop("values", None)
    out["field_values"] = pretty
    payload = json.dumps(out, ensure_ascii=False, indent=2).encode("utf-8")
    filename = f"form_record_{record_id}.json"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(io.BytesIO(payload), media_type="application/json; charset=utf-8", headers=headers)


@router.get("/{record_id}/export/csv")
def export_csv(
    record_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    rec = db.query(FormRecord).filter(FormRecord.id == record_id).first()
    if not rec:
        raise HTTPException(status_code=404, detail="表单记录不存在")
    tpl = get_template_or_404(db, rec.template_id)
    fields = sorted_fields_from_template(tpl)
    values_by_key = {v.field_key: v.field_value for v in rec.values}
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["表单记录ID", rec.id])
    w.writerow(["标题", rec.title])
    w.writerow(["模板", tpl.name])
    w.writerow(["状态", rec.status])
    w.writerow(["机台ID", rec.equipment_id or ""])
    w.writerow(["批号", rec.batch_no or ""])
    w.writerow(["班次", rec.shift or ""])
    w.writerow(["生产日期", (str(rec.production_date)[:10]) if rec.production_date else ""])
    w.writerow(["填写人ID", rec.filled_by or ""])
    w.writerow(["提交时间", str(rec.submitted_at) if rec.submitted_at else ""])
    w.writerow(["创建时间", str(rec.created_at)])
    w.writerow([])
    w.writerow(["序号", "字段Key", "字段标签", "字段类型", "单位", "填写值"])
    for idx, f in enumerate(fields, start=1):
        v = values_by_key.get(f["key"])
        # 复杂类型 json 化
        if isinstance(v, (dict, list)):
            v_s = json.dumps(v, ensure_ascii=False)
        elif v is None:
            v_s = ""
        else:
            v_s = str(v)
        unit = f.get("unit") or ""
        w.writerow([idx, f["key"], f.get("label") or f["key"], f["type"], unit, v_s])
    if rec.remark:
        w.writerow([])
        w.writerow(["备注", rec.remark])
    filename = f"form_record_{record_id}.csv"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    # BOM 便于 Excel 打开中文不乱码
    body = ("\ufeff" + buf.getvalue()).encode("utf-8-sig")
    return StreamingResponse(io.BytesIO(body), media_type="text/csv; charset=utf-8-sig", headers=headers)


# ============================================================
# 导出 Excel（openpyxl，单条 + 批量）
# ============================================================

def _value_to_cell(v):
    """将字段值转为 Excel 单元格友好类型。"""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "是" if v else "否"
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False)
    return v


@router.get(
    "/{record_id}/export/excel",
    dependencies=[Depends(require_permission("production.process_export"))],
)
def export_excel_single(
    record_id: int,
    db: Session = Depends(get_db),
):
    """单条记录导出为 Excel（沿用 CSV 的纵向布局）。"""
    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="服务端未安装 openpyxl，无法导出 Excel")

    rec = db.query(FormRecord).filter(FormRecord.id == record_id).first()
    if not rec:
        raise HTTPException(status_code=404, detail="表单记录不存在")
    tpl = get_template_or_404(db, rec.template_id)
    fields = sorted_fields_from_template(tpl)
    values_by_key = {v.field_key: v.field_value for v in rec.values}

    wb = Workbook()
    ws = wb.active
    ws.title = "工艺记录"
    meta = [
        ("表单记录ID", rec.id),
        ("标题", rec.title),
        ("模板", tpl.name),
        ("状态", rec.status),
        ("机台ID", rec.equipment_id or ""),
        ("批号", rec.batch_no or ""),
        ("班次", rec.shift or ""),
        ("生产日期", (str(rec.production_date)[:10]) if rec.production_date else ""),
        ("填写人ID", rec.filled_by or ""),
        ("提交时间", str(rec.submitted_at) if rec.submitted_at else ""),
        ("创建时间", str(rec.created_at)),
    ]
    for r_idx, (k, v) in enumerate(meta, start=1):
        ws.cell(row=r_idx, column=1, value=k)
        ws.cell(row=r_idx, column=2, value=v)
    start = len(meta) + 2
    ws.cell(row=start, column=1, value="序号")
    ws.cell(row=start, column=2, value="字段Key")
    ws.cell(row=start, column=3, value="字段标签")
    ws.cell(row=start, column=4, value="字段类型")
    ws.cell(row=start, column=5, value="单位")
    ws.cell(row=start, column=6, value="填写值")
    for col in range(1, 7):
        ws.cell(row=start, column=col).font = ws.cell(row=start, column=col).font.copy(bold=True)
    for i, f in enumerate(fields, start=1):
        ws.cell(row=start + i, column=1, value=i)
        ws.cell(row=start + i, column=2, value=f["key"])
        ws.cell(row=start + i, column=3, value=f.get("label") or f["key"])
        ws.cell(row=start + i, column=4, value=f["type"])
        ws.cell(row=start + i, column=5, value=f.get("unit") or "")
        ws.cell(row=start + i, column=6, value=_value_to_cell(values_by_key.get(f["key"])))
    # 列宽自适应（粗略）
    for col_idx, w in enumerate([8, 24, 24, 12, 10, 40], start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w

    if rec.remark:
        ws.cell(row=start + len(fields) + 2, column=1, value="备注")
        ws.cell(row=start + len(fields) + 2, column=2, value=rec.remark)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"form_record_{record_id}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.get(
    "/export/excel",
    dependencies=[Depends(require_permission("production.process_export"))],
)
def export_excel_bulk(
    template_id: Optional[int] = Query(None, description="按模板筛选（推荐先选模板，列结构按模板字段展开）"),
    equipment_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    batch_no: Optional[str] = Query(None),
    shift: Optional[str] = Query(None),
    production_date_from: Optional[str] = Query(None),
    production_date_to: Optional[str] = Query(None),
    limit: int = Query(500, ge=1, le=5000),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """批量导出工艺记录为 Excel：每行一条记录，列 = 元信息 + 模板字段。

    若 template_id 指定，则按该模板的字段 schema 展开列；
    若未指定，则按所有记录的并集字段展开（先按 field_key 排序）。
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="服务端未安装 openpyxl，无法导出 Excel")

    q = db.query(FormRecord)
    if template_id:
        q = q.filter(FormRecord.template_id == template_id)
    if equipment_id:
        q = q.filter(FormRecord.equipment_id == equipment_id)
    if status:
        q = q.filter(FormRecord.status == status)
    if batch_no:
        q = q.filter(FormRecord.batch_no.ilike(f"%{batch_no}%"))
    if shift:
        q = q.filter(FormRecord.shift == shift)
    if production_date_from:
        try:
            dt_from = datetime.fromisoformat(production_date_from.replace("Z", ""))
            q = q.filter(FormRecord.production_date >= dt_from)
        except ValueError:
            pass
    if production_date_to:
        try:
            dt_to = datetime.fromisoformat(production_date_to.replace("Z", ""))
            # 仅传日期(如 2026-08-20)时 fromisoformat 解析为当天 00:00，
            # 会漏掉当天 00:00 之后的记录；补到当天 23:59:59 兜底。
            if len(production_date_to) == 10:
                dt_to = dt_to.replace(hour=23, minute=59, second=59)
            q = q.filter(FormRecord.production_date <= dt_to)
        except ValueError:
            pass
    records = q.order_by(FormRecord.created_at.desc()).limit(limit).all()
    if not records:
        raise HTTPException(status_code=404, detail="按条件未匹配到任何表单记录")

    # 收集字段定义
    tpl_cache: dict[int, tuple] = {}  # tpl_id -> (tpl, sorted_fields)

    def get_tpl(tpl_id: int):
        if tpl_id not in tpl_cache:
            tpl = db.query(FormTemplate).filter(FormTemplate.id == tpl_id).first()
            if tpl:
                tpl_cache[tpl_id] = (tpl, sorted_fields_from_template(tpl))
            else:
                tpl_cache[tpl_id] = (None, [])
        return tpl_cache[tpl_id]

    if template_id:
        tpl, fields = get_tpl(template_id)
        field_keys = [f["key"] for f in fields]
        field_labels = [(f.get("label") or f["key"]) for f in fields]
    else:
        # 取并集
        seen = {}
        for r in records:
            _, fields = get_tpl(r.template_id)
            for f in fields:
                seen.setdefault(f["key"], f.get("label") or f["key"])
        field_keys = list(seen.keys())
        field_labels = [seen[k] for k in field_keys]

    meta_cols = ["记录ID", "标题", "模板", "状态", "机台ID", "机台名", "批号", "班次",
                 "生产日期", "填写人ID", "提交时间", "创建时间"]

    wb = Workbook()
    ws = wb.active
    ws.title = "工艺数据"
    header = meta_cols + field_labels
    for col_idx, h in enumerate(header, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = c.font.copy(bold=True)

    for row_idx, r in enumerate(records, start=2):
        tpl, _ = get_tpl(r.template_id)
        eq_name = None
        if r.equipment_id:
            eq = db.query(Equipment).filter(Equipment.id == r.equipment_id).first()
            eq_name = eq.name if eq else None
        meta_row = [
            r.id, r.title, tpl.name if tpl else "", r.status,
            r.equipment_id or "", eq_name or "",
            r.batch_no or "", r.shift or "",
            (str(r.production_date)[:10]) if r.production_date else "",
            r.filled_by or "",
            str(r.submitted_at) if r.submitted_at else "",
            str(r.created_at),
        ]
        values_by_key = {v.field_key: v.field_value for v in r.values}
        field_row = [_value_to_cell(values_by_key.get(k)) for k in field_keys]
        for col_idx, v in enumerate(meta_row + field_row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=v)

    # 列宽粗调
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(40, len(header[col_idx - 1]) * 2 + 6))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"form_records_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{fname}"'}
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
